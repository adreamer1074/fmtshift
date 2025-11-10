"""
Excel ライター - 中間形式からExcelファイルを生成
"""
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from converters.base import Document, Sheet, Content, ContentType, Table


class ExcelWriter:
    """中間形式からExcelファイルを生成"""
    
    def write(self, document: Document, file_path: Path):
        """
        Documentオブジェクトをアelファイルに書き込み
        
        Args:
            document: 中間形式のドキュメント
            file_path: 出力先Excelファイルのパス
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # デフォルトシートを削除
        
        for sheet_data in document.sheets:
            ws = self._create_sheet(wb, sheet_data.name)
            self._write_sheet_content(ws, sheet_data)
        
        # シートが一つもない場合はデフォルトを追加
        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet('Sheet1')
            ws.cell(row=1, column=1, value='(空のドキュメント)')
        
        wb.save(file_path)
    
    def _create_sheet(self, wb, sheet_name: str):
        """有効なシート名でシートを作成"""
        # シート名を有効な名前に変換（Excelの制約に対応）
        valid_name = sheet_name[:31]  # 最大31文字
        # 無効な文字を削除
        for char in ['\\', '/', '?', '*', '[', ']', ':']:
            valid_name = valid_name.replace(char, '')
        
        # 同名シートが存在する場合は番号を付ける
        if valid_name in wb.sheetnames:
            counter = 1
            while f'{valid_name}_{counter}' in wb.sheetnames:
                counter += 1
            valid_name = f'{valid_name}_{counter}'
        
        return wb.create_sheet(valid_name)
    
    def _write_sheet_content(self, ws, sheet: Sheet):
        """シートにコンテンツを書き込み"""
        row_idx = 1
        table_start_row = None
        
        for content in sheet.contents:
            if content.type == ContentType.TITLE:
                # タイトル（太字、大きめ）
                cell = ws.cell(row=row_idx, column=1, value=content.value)
                cell.font = Font(size=14, bold=True)
                row_idx += 1
            
            elif content.type == ContentType.TEXT:
                # 通常のテキスト
                ws.cell(row=row_idx, column=1, value=content.value)
                row_idx += 1
            
            elif content.type == ContentType.LIST_ITEM:
                # リスト項目
                ws.cell(row=row_idx, column=1, value=f'• {content.value}')
                row_idx += 1
            
            elif content.type == ContentType.NUMBERED_LIST:
                # 番号付きリスト
                ws.cell(row=row_idx, column=1, value=content.value)
                row_idx += 1
            
            elif content.type == ContentType.CODE_BLOCK:
                # コードブロック
                for line in content.value.split('\n'):
                    cell = ws.cell(row=row_idx, column=1, value=line)
                    cell.font = Font(name='Courier New', size=9)
                    row_idx += 1
            
            elif content.type == ContentType.TABLE:
                # テーブル
                table: Table = content.value
                
                # ヘッダー行
                for col_idx, header in enumerate(table.headers, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                row_idx += 1
                
                # データ行
                for row_data in table.rows:
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    row_idx += 1
            
            elif content.type == ContentType.EMPTY:
                # 空行
                row_idx += 1
        
        # 列幅を自動調整
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, ws):
        """列幅を自動調整"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)  # 最大80
            ws.column_dimensions[column_letter].width = adjusted_width
