"""
Excel パーサー - Excelファイルを中間形式に変換
"""
from pathlib import Path
import openpyxl
from converters.base import Document, Sheet, Content, ContentType, Table


class ExcelParser:
    """Excelファイルを読み込んで中間形式に変換"""
    
    def parse(self, file_path: Path) -> Document:
        """
        Excelファイルを解析してDocumentオブジェクトに変換
        
        Args:
            file_path: Excelファイルのパス
            
        Returns:
            Document: 中間形式のドキュメント
        """
        wb = openpyxl.load_workbook(file_path)
        doc = Document(title=file_path.stem)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet = Sheet(name=sheet_name)
            
            # シートからデータを抽出
            data = []
            for row in ws.iter_rows(values_only=True):
                # 空行をスキップ
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                # None を空文字列に変換
                data.append([str(cell) if cell is not None else '' for cell in row])
            
            if data:
                # テーブルとして処理（1行目をヘッダー）
                if len(data) > 1:
                    headers = data[0]
                    rows = data[1:]
                    table = Table(headers=headers, rows=rows)
                    content = Content(
                        type=ContentType.TABLE,
                        value=table,
                        metadata={'source': 'excel'}
                    )
                    sheet.add_content(content)
                elif len(data) == 1:
                    # ヘッダーのみの場合
                    table = Table(headers=data[0], rows=[])
                    content = Content(
                        type=ContentType.TABLE,
                        value=table,
                        metadata={'source': 'excel'}
                    )
                    sheet.add_content(content)
            
            doc.add_sheet(sheet)
        
        return doc
